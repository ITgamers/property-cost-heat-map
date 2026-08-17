#!/usr/bin/env python3
"""
Build-time ETL for the Property Cost Heat Map.

Fetches authoritative public data and bakes it into static JSON/GeoJSON that the
web app consumes. Nothing is fetched at page-load time, so the map works offline
and never trips over government CORS policies.

Sources
-------
  Tax rates    Texas Comptroller "Rates and Levies" XLSX (annual, certified ~Oct)
               https://comptroller.texas.gov/taxes/property-tax/rates/
  Boundaries   US Census TIGERweb GeoServices REST API (school districts,
               incorporated places, counties)
  Mortgage     FRED MORTGAGE30US - Freddie Mac Primary Mortgage Market Survey
               (weekly, keyless CSV endpoint)

Usage
-----
  python build_data.py                 # incremental, uses cached downloads
  python build_data.py --refresh       # re-download everything
  python build_data.py --year 2025     # target a different tax year

Output lands in ../web/data/.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path

import requests
from shapely.geometry import shape, mapping
from shapely.ops import unary_union
from shapely.validation import make_valid

HERE = Path(__file__).parent
RAW = HERE / "raw"
OUT = HERE.parent / "web" / "data"

COMPTROLLER = "https://comptroller.texas.gov/taxes/property-tax/docs"
TIGER = "https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb"
FRED_CSV = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}"

# --- Study area -------------------------------------------------------------
# San Antonio-New Braunfels MSA. The engine is state-agnostic; this list is the
# only thing that scopes v1 to San Antonio. Add counties to widen coverage.
STUDY_COUNTIES = {
    "Bexar": "029",
    "Comal": "091",
    "Guadalupe": "187",
    "Kendall": "259",
    "Medina": "325",
    "Wilson": "493",
    "Atascosa": "013",
}
STATE_FIPS = "48"  # Texas
BBOX = (-99.05, 28.95, -97.85, 30.15)  # lon/lat envelope covering the MSA

# Simplification tolerance in degrees (~55 m). Keeps the payload small enough to
# load instantly while staying visually faithful at metro zoom levels.
SIMPLIFY_TOL = 0.0005
# Intersections smaller than this (sq. degrees, ~1.1 sq km) are topological
# slivers from boundary digitisation, not real places.
MIN_ZONE_AREA = 1e-4

# --- Taxing unit type codes (position 3 of the Comptroller TAXING UNIT ID) ---
# These drive the rules engine's classification of each levy.
UNIT_TYPES = {
    "02": "school",
    "03": "city",
    "04": "mud",
    "07": "lid",
    "08": "mud",
    "09": "sid",  # special/public improvement district (city-linked)
    "11": "hospital",
    "13": "wcid",
    "15": "college",
    "19": "wcid",
    "22": "flood",
    "27": "river",
    "40": "esd",
}
# Levied on every parcel in the county, regardless of city/school district.
COUNTYWIDE_TYPES = {"hospital", "college", "river", "flood"}
# Parcel-specific; cannot be geolocated from these sources. Surfaced in the UI
# as opt-in toggles rather than silently applied or silently omitted.
OPTIONAL_TYPES = {"mud", "wcid", "lid", "esd"}


# ---------------------------------------------------------------------------
# Fetch helpers
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    print(f"  {msg}", flush=True)


def write_pair(stem: str, payload, global_name: str) -> None:
    """
    Emit both a plain data file and a JS file that assigns it to a global.

    The .geojson/.json is the portable artifact (reusable in QGIS, other tools).
    The .js twin lets index.html open straight off the filesystem - browsers
    block fetch() over file://, so a <script> tag is the only dependency-free
    way to make double-clicking the page work.
    """
    ext = "geojson" if isinstance(payload, dict) and payload.get("type") == "FeatureCollection" else "json"
    blob = json.dumps(payload, separators=(",", ":"))
    (OUT / f"{stem}.{ext}").write_text(
        blob if ext == "geojson" else json.dumps(payload, indent=2), encoding="utf-8"
    )
    (OUT / f"{stem}.js").write_text(
        f"window.{global_name}={blob};", encoding="utf-8"
    )


def download(url: str, dest: Path, refresh: bool = False) -> Path:
    """Download to dest, reusing the cached copy unless refresh is set."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and not refresh:
        log(f"cached  {dest.name}")
        return dest
    log(f"fetch   {url}")
    r = requests.get(url, timeout=180)
    r.raise_for_status()
    dest.write_bytes(r.content)
    return dest


def tiger_query(service: str, layer: int, where: str, out_fields: str) -> list[dict]:
    """
    Query a TIGERweb layer and return GeoJSON features.

    TIGERweb caps result counts, so we page with resultOffset until exhausted.
    """
    features: list[dict] = []
    offset = 0
    while True:
        params = {
            "where": where,
            "geometry": f"{BBOX[0]},{BBOX[1]},{BBOX[2]},{BBOX[3]}",
            "geometryType": "esriGeometryEnvelope",
            "inSR": "4326",
            "outSR": "4326",
            "spatialRel": "esriSpatialRelIntersects",
            "outFields": out_fields,
            "returnGeometry": "true",
            "f": "geojson",
            "resultOffset": offset,
        }
        for attempt in range(3):
            try:
                r = requests.get(
                    f"{TIGER}/{service}/MapServer/{layer}/query",
                    params=params,
                    timeout=180,
                )
                r.raise_for_status()
                payload = r.json()
                break
            except Exception as exc:  # transient census outage
                if attempt == 2:
                    raise
                log(f"retry   {exc.__class__.__name__}, attempt {attempt + 2}/3")
                time.sleep(3)

        batch = payload.get("features", [])
        features.extend(batch)
        if not payload.get("exceededTransferLimit") or not batch:
            break
        offset += len(batch)
    return features


# ---------------------------------------------------------------------------
# Comptroller rate parsing
# ---------------------------------------------------------------------------

@dataclass
class TaxUnit:
    unit_id: str
    name: str
    county: str
    kind: str
    rate: float
    mo_rate: float = 0.0
    is_rate: float = 0.0


def _header_row(ws):
    """Header position varies between the four workbooks; find it by marker."""
    for row in ws.iter_rows(max_row=12, values_only=True):
        if row and row[0] == "CAD ID":
            return list(row)
    raise ValueError("no header row found")


def parse_rates(path: Path, kind_hint: str | None = None) -> list[TaxUnit]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = None
    for row in rows:
        if row and row[0] == "CAD ID":
            header = list(row)
            break
    if header is None:
        raise ValueError(f"{path.name}: no header row")

    def col(*candidates: str) -> int | None:
        for c in candidates:
            if c in header:
                return header.index(c)
        return None

    i_county = col("COUNTY NAME")
    i_name = col("TAXING UNIT NAME", "TU NAME")
    i_id = col("TAXING UNIT ID")
    i_rate = col("TOTAL TAX RATE", "TOTAL COUNTY TAX RATE")
    i_mo = col("M&O RATE", "GF M&O TAX RATE")
    i_is = col("I&S RATE", "GF I&S TAX RATE")

    units: list[TaxUnit] = []
    for row in rows:
        if not row or i_county is None or row[i_county] not in STUDY_COUNTIES:
            continue
        try:
            rate = float(row[i_rate] or 0)
        except (TypeError, ValueError):
            continue

        unit_id = str(row[i_id]).strip() if i_id is not None else ""
        if kind_hint:
            kind = kind_hint
        else:
            parts = unit_id.split("-")
            kind = UNIT_TYPES.get(parts[2], "other") if len(parts) > 2 else "other"

        name = (
            str(row[i_name]).strip()
            if i_name is not None and row[i_name]
            else f"{row[i_county]} County"
        )
        # Comptroller flags provisional/partial figures with trailing asterisks.
        name = name.rstrip("* ").strip()

        def num(idx):
            try:
                return float(row[idx] or 0)
            except (TypeError, ValueError):
                return 0.0

        units.append(
            TaxUnit(
                unit_id=unit_id,
                name=name,
                county=str(row[i_county]),
                kind=kind,
                rate=rate,
                mo_rate=num(i_mo) if i_mo is not None else 0.0,
                is_rate=num(i_is) if i_is is not None else 0.0,
            )
        )
    wb.close()
    return units


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def clean(geom):
    """Repair self-intersections that TIGER polygons occasionally carry."""
    if not geom.is_valid:
        geom = make_valid(geom)
    return geom


def norm_isd(name: str) -> str:
    """'Northside Independent School District' -> 'Northside ISD' (Comptroller form)."""
    return (
        name.replace("Independent School District", "ISD")
        .replace("Consolidated ISD", "CISD")
        .strip()
    )


def build_zones(isds, cities, counties, rates_by_county) -> list[dict]:
    """
    Intersect school districts x incorporated places x counties.

    School district drives roughly half the levy and its boundaries ignore city
    and county lines, so the intersection - not the county - is the real unit of
    cost geography. Each output zone is an area where every homeowner faces an
    identical mandatory rate stack.
    """
    zones: list[dict] = []

    for county_name, county_geom in counties.items():
        county_units = rates_by_county.get(county_name, {})
        county_rate = county_units.get("county", 0.0)
        countywide = county_units.get("countywide", [])
        countywide_rate = sum(u["rate"] for u in countywide)

        for isd_name, isd_geom in isds.items():
            if not isd_geom.intersects(county_geom):
                continue
            isd_in_county = clean(isd_geom.intersection(county_geom))
            if isd_in_county.is_empty or isd_in_county.area < MIN_ZONE_AREA:
                continue

            isd_rate = county_units.get("schools", {}).get(norm_isd(isd_name))
            if isd_rate is None:
                # District overlaps the county line but levies through its home
                # county's appraisal roll; skip rather than invent a rate.
                continue

            # Carve out each incorporated place, then keep the remainder as the
            # unincorporated balance (which pays no city tax but often an ESD).
            remainder = isd_in_county
            for city_name, city_geom in cities.items():
                if not city_geom.intersects(isd_in_county):
                    continue
                piece = clean(isd_in_county.intersection(city_geom))
                if piece.is_empty or piece.area < MIN_ZONE_AREA:
                    continue
                city_rate = county_units.get("cities", {}).get(city_name)
                if city_rate is None:
                    continue
                zones.append(
                    _zone(piece, county_name, county_rate, countywide, countywide_rate,
                          isd_name, isd_rate, city_name, city_rate)
                )
                remainder = clean(remainder.difference(city_geom))

            if not remainder.is_empty and remainder.area >= MIN_ZONE_AREA:
                zones.append(
                    _zone(remainder, county_name, county_rate, countywide,
                          countywide_rate, isd_name, isd_rate, None, 0.0)
                )
    return zones


def _zone(geom, county, county_rate, countywide, countywide_rate,
          isd, isd_rate, city, city_rate) -> dict:
    total = county_rate + countywide_rate + isd_rate + city_rate
    label = f"{isd} / {city or 'Unincorporated'} / {county} County"
    return {
        "type": "Feature",
        "geometry": mapping(geom.simplify(SIMPLIFY_TOL, preserve_topology=True)),
        "properties": {
            "zone_id": label.lower().replace(" ", "-").replace("/", ""),
            "label": label,
            "county": county,
            "city": city,
            "isd": isd,
            "rates": {
                "county": round(county_rate, 6),
                "city": round(city_rate, 6),
                "school": round(isd_rate, 6),
                "countywide": [
                    {"name": u["name"], "rate": round(u["rate"], 6)}
                    for u in countywide
                ],
            },
            "total_rate": round(total, 6),
        },
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--year", default="2025", help="tax year to pull (default 2025)")
    ap.add_argument("--refresh", action="store_true", help="re-download sources")
    ap.add_argument(
        "--repo",
        default="ITgamers/property-cost-heat-map",
        help="GitHub repo the app pulls refreshed data from (owner/name). "
             "Change this if you fork.",
    )
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)
    provenance: list[dict] = []

    # 1. Tax rates -----------------------------------------------------------
    print("\n[1/4] Texas Comptroller rates and levies")
    files = {
        "county": f"{args.year}-county-rates-levies.xlsx",
        "school": f"{args.year}-school-district-rates-levies.xlsx",
        "city": f"{args.year}-city-rates-levies.xlsx",
        "special": f"{args.year}-special-district-rates-levies.xlsx",
    }
    units: list[TaxUnit] = []
    for kind, fname in files.items():
        path = download(f"{COMPTROLLER}/{fname}", RAW / fname, args.refresh)
        hint = {"county": "county", "school": "school", "city": "city"}.get(kind)
        units.extend(parse_rates(path, hint))
        provenance.append({"source": "Texas Comptroller", "file": fname,
                           "url": f"{COMPTROLLER}/{fname}"})
    log(f"parsed {len(units)} taxing units across {len(STUDY_COUNTIES)} counties")

    rates_by_county: dict[str, dict] = {}
    optional_by_county: dict[str, list] = {}
    for u in units:
        b = rates_by_county.setdefault(
            u.county, {"county": 0.0, "schools": {}, "cities": {}, "countywide": []}
        )
        if u.kind == "county":
            b["county"] = u.rate
        elif u.kind == "school":
            b["schools"][u.name] = u.rate
        elif u.kind == "city":
            b["cities"][u.name] = u.rate
        elif u.kind in COUNTYWIDE_TYPES:
            b["countywide"].append({"name": u.name, "rate": u.rate, "kind": u.kind})
        elif u.kind in OPTIONAL_TYPES and u.rate > 0:
            optional_by_county.setdefault(u.county, []).append(
                {"name": u.name, "rate": round(u.rate, 6), "kind": u.kind}
            )
    for lst in optional_by_county.values():
        lst.sort(key=lambda d: -d["rate"])

    # 2. Boundaries ----------------------------------------------------------
    print("\n[2/4] Census TIGERweb boundaries")
    county_where = " OR ".join(
        f"(STATE='{STATE_FIPS}' AND COUNTY='{fips}')" for fips in STUDY_COUNTIES.values()
    )
    raw_counties = tiger_query(
        "State_County", 1, county_where, "GEOID,NAME,BASENAME,STATE,COUNTY"
    )
    # TIGER says "Bexar County"; the Comptroller says "Bexar". BASENAME is the
    # bare form and is what the rate tables key on.
    counties = {
        f["properties"]["BASENAME"]: clean(shape(f["geometry"]))
        for f in raw_counties
        if f.get("geometry")
    }
    log(f"{len(counties)} counties")

    raw_isds = tiger_query("School", 0, f"STATE='{STATE_FIPS}'", "GEOID,NAME,STATE")
    isds = {
        f["properties"]["NAME"]: clean(shape(f["geometry"]))
        for f in raw_isds
        if f.get("geometry")
    }
    log(f"{len(isds)} school districts")

    raw_cities = tiger_query(
        "Places_CouSub_ConCity_SubMCD", 4, f"STATE='{STATE_FIPS}'", "GEOID,NAME,BASENAME,STATE"
    )
    cities = {
        f["properties"].get("BASENAME") or f["properties"]["NAME"]: clean(shape(f["geometry"]))
        for f in raw_cities
        if f.get("geometry")
    }
    log(f"{len(cities)} incorporated places")
    provenance.append({"source": "US Census TIGERweb", "url": TIGER,
                       "layers": ["School/0", "Places_CouSub_ConCity_SubMCD/4",
                                  "State_County/1"]})

    # 3. Intersect -----------------------------------------------------------
    print("\n[3/4] Intersecting school district x city x county")
    zones = build_zones(isds, cities, counties, rates_by_county)
    zones.sort(key=lambda z: z["properties"]["total_rate"])
    log(f"{len(zones)} distinct tax zones")
    if zones:
        lo, hi = zones[0]["properties"], zones[-1]["properties"]
        log(f"lowest   {lo['total_rate']:.4f}%  {lo['label']}")
        log(f"highest  {hi['total_rate']:.4f}%  {hi['label']}")

    fc = {"type": "FeatureCollection", "features": zones}
    write_pair("zones", fc, "ZONES")
    log(f"wrote zones.geojson ({(OUT / 'zones.geojson').stat().st_size / 1e6:.1f} MB)")

    # 4. Market data ---------------------------------------------------------
    print("\n[4/4] Market inputs")
    mortgage_rate, mortgage_date = 6.67, "unknown"
    try:
        r = requests.get(FRED_CSV.format(series="MORTGAGE30US"), timeout=60)
        r.raise_for_status()
        rows = [
            row for row in csv.reader(io.StringIO(r.text))
            if len(row) == 2 and row[1] not in ("", ".")
        ][1:]
        mortgage_date, mortgage_rate = rows[-1][0], float(rows[-1][1])
        log(f"30-yr fixed {mortgage_rate}% as of {mortgage_date} (Freddie Mac PMMS)")
        provenance.append({"source": "FRED / Freddie Mac PMMS",
                           "series": "MORTGAGE30US",
                           "url": "https://fred.stlouisfed.org/series/MORTGAGE30US"})
    except Exception as exc:
        log(f"FRED unavailable ({exc}); falling back to {mortgage_rate}%")

    market = {
        "generated": time.strftime("%Y-%m-%d"),
        "tax_year": args.year,
        "mortgage": {
            "rate_30yr": mortgage_rate,
            # PMMS publishes only the 30-yr and 15-yr headline. Spreads below are
            # conventional market rules of thumb, adjustable in the UI.
            "spread_15yr": -0.75,
            "spread_fha": -0.25,
            "spread_va": -0.30,
            "as_of": mortgage_date,
            "source": "Freddie Mac PMMS via FRED (MORTGAGE30US)",
        },
        # No free per-ZIP homeowners insurance API exists; the good data
        # (Quadrant/Insure.com) is proprietary. This is a transparent model:
        # statewide average scaled by dwelling value, with a county hail/risk
        # multiplier. Always user-overridable and labelled as an estimate.
        "insurance": {
            "model": "value_scaled",
            "state_avg_annual": 4350,
            "state_avg_dwelling": 300000,
            "min_annual": 900,
            "county_factor": {
                "Bexar": 1.00, "Comal": 1.02, "Guadalupe": 1.01, "Kendall": 1.03,
                "Medina": 0.98, "Wilson": 0.97, "Atascosa": 0.96,
            },
            "source": "Insure.com 2026 Texas average ($4,350 @ $300k dwelling)",
            "note": "Modelled estimate, not a quote. Override with a real quote.",
        },
        # Texas exemption rules for the rules engine. Keyed so other states can
        # be added as sibling rule sets without touching engine code.
        "exemptions": {
            "TX": {
                "school_homestead": 140000,
                "school_homestead_senior_extra": 60000,
                "local_optional_pct": 0.20,
                "local_optional_applies_to": ["county", "city", "countywide"],
                "senior_flat": 3000,
                "appraisal_cap_pct": 0.10,
                "note": "Prop 13 (2025) raised the school homestead exemption to "
                        "$140,000 effective 2026-01-01. Bexar County and the City "
                        "of San Antonio each grant the maximum 20% local option.",
            }
        },
        "optional_districts": optional_by_county,
        "provenance": provenance,
        # Where the in-app "Update data" button looks for fresher output.
        # Neither FRED nor the Comptroller sends CORS headers, so a browser
        # cannot refresh from them directly. A scheduled GitHub Action re-runs
        # this script and commits the result; raw.githubusercontent.com serves
        # it with Access-Control-Allow-Origin: *, which any origin can read -
        # including a page opened straight off the filesystem.
        "update": {
            "base_url": f"https://raw.githubusercontent.com/{args.repo}/main/web/data",
            "repo": args.repo,
        },
    }
    write_pair("market", market, "MARKET")
    log("wrote market.json")

    print(f"\nDone. Output in {OUT}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
